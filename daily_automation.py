import datetime
import zipfile
import io
import os
import gc
from io import BytesIO

import pandas as pd
import numpy as np
import openpyxl 
from openpyxl.utils.dataframe import dataframe_to_rows 
import xlsxwriter 

# Google API Libraries
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
import google.auth

# --- LOGGING UTILITY ---

def log(message):
    """Prints a message with a timestamp for debugging/logs."""
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(f"[{timestamp}] {message}")

# --- AUTHENTICATION & CONFIGURATION ---
SERVICE_ACCOUNT_FILE = 'credentials.json' 
SCOPES = ['https://www.googleapis.com/auth/drive']

# Folder IDs
SOURCE_FOLDER_ID = '1sern1xXqdDrQQBLXxbANj7LPs3IE1Dzo'
TARGET_FOLDER_ID = '1HTkBss1orVVn1akNygxiuklaAHQnJy8w'
FILE_PREFIXES = ['ArticleSalesReport', 'Overall_Instock']

def authenticate():
    """Authenticates using the Service Account and returns the Drive Service."""
    log("Attempting Google Drive authentication...")
    try:
        creds, _ = google.auth.default(scopes=SCOPES)
        drive_service = build('drive', 'v3', credentials=creds)
        log("✅ Google Drive authentication successful.")
        return drive_service
    except Exception as e:
        log(f"❌ ERROR: Authentication failed. Details: {e}")
        return None

# ==============================================================================
# SECTION 1: HELPER FUNCTIONS (DOWNLOAD / UPLOAD)
# ==============================================================================

def download_csv_to_df(drive_service, file_name, folder_id):
    """Finds a CSV helper file by name and loads it into a Pandas DataFrame."""
    log(f"  Downloading helper file: {file_name}...")
    try:
        query = f"'{folder_id}' in parents and name='{file_name}' and trashed=false"
        results = drive_service.files().list(
            q=query, fields="files(id, name)", supportsAllDrives=True, includeItemsFromAllDrives=True
        ).execute()
        items = results.get('files', [])
        
        if not items:
            log(f"  [ERROR] Helper file '{file_name}' not found.")
            return None
            
        request = drive_service.files().get_media(fileId=items[0]['id'], supportsAllDrives=True)
        file_buffer = BytesIO()
        downloader = MediaIoBaseDownload(file_buffer, request)
        
        done = False
        while not done: _, done = downloader.next_chunk()
        
        file_buffer.seek(0)
        try:
            return pd.read_csv(file_buffer)
        except UnicodeDecodeError:
            file_buffer.seek(0)
            log(f"  [INFO] Reading '{file_name}' with latin1 encoding.")
            return pd.read_csv(file_buffer, encoding='latin1')

    except Exception as e:
        log(f"  [ERROR] Failed to download '{file_name}': {e}")
        return None

def load_file_to_df(drive_service, file_id, file_name):
    """Downloads a generic file (CSV or ZIP containing CSV) into a DataFrame."""
    log(f"  Loading main file: {file_name} (ID: {file_id})...")
    try:
        request = drive_service.files().get_media(fileId=file_id, supportsAllDrives=True)
        file_buffer = BytesIO()
        downloader = MediaIoBaseDownload(file_buffer, request)
        
        done = False
        while not done: _, done = downloader.next_chunk()
        file_buffer.seek(0)

        csv_bytes = None
        if file_name.endswith('.zip'):
            log(f"    > Unzipping {file_name}...")
            with zipfile.ZipFile(file_buffer, 'r') as zf:
                csv_names = [n for n in zf.namelist() if n.endswith('.csv') and not n.startswith('__MACOSX')]
                if not csv_names:
                    log("    [ERROR] No CSV found inside zip.")
                    return None
                csv_bytes = zf.read(csv_names[0])
        elif file_name.endswith('.csv'):
            csv_bytes = file_buffer.read()
        else:
            log(f"    [ERROR] Unsupported file format: {file_name}")
            return None

        if csv_bytes:
            try:
                # Read everything as string initially to prevent auto-conversion errors
                return pd.read_csv(BytesIO(csv_bytes), low_memory=False)
            except UnicodeDecodeError:
                return pd.read_csv(BytesIO(csv_bytes), encoding='latin1', low_memory=False)
        return None

    except Exception as e:
        log(f"  [ERROR] Failed to load main file: {e}")
        return None

def upload_excel_report(drive_service, excel_buffer, file_name, folder_id):
    """Uploads the generated Excel Report (with graphs) to Drive."""
    log(f"  Upload: {file_name}...")
    try:
        query = f"'{folder_id}' in parents and name='{file_name}' and trashed=false"
        results = drive_service.files().list(q=query, fields="files(id)").execute()
        existing_files = results.get('files', [])

        file_metadata = {'name': file_name, 'parents': [folder_id]}
        media = MediaIoBaseUpload(
            excel_buffer, 
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
            resumable=True
        )
        
        if existing_files:
            file_id = existing_files[0]['id']
            drive_service.files().update(fileId=file_id, media_body=media, supportsAllDrives=True).execute()
        else:
            drive_service.files().create(body=file_metadata, media_body=media, fields='id', supportsAllDrives=True).execute()
            
        log(f"  ✅ Advanced Excel Report uploaded successfully.")
    except Exception as e:
        log(f"  ❌ Failed to upload report: {e}")

def update_xlsm_data_sheet(drive_service, df_to_paste, file_name_to_find, sheet_name_to_update, folder_id):
    """
    Updates the raw data sheet in the macro-enabled .xlsm file.
    CRITICAL SECURITY: Sets the data sheet to 'veryHidden' so users cannot see data
    unless they log in via the VBA Macro.
    """
    log(f"\n--- Updating Raw Data in Excel: {file_name_to_find} ---")
    if df_to_paste is None: return

    try:
        # 1. Download the existing XLSM file
        query = f"'{folder_id}' in parents and name='{file_name_to_find}' and trashed=false"
        results = drive_service.files().list(q=query, fields="files(id, name)", supportsAllDrives=True, includeItemsFromAllDrives=True).execute()
        items = results.get('files', [])
        if not items:
            log(f"  [ERROR] {file_name_to_find} not found.")
            return
        
        file_id = items[0]['id']
        request = drive_service.files().get_media(fileId=file_id, supportsAllDrives=True)
        buffer = BytesIO()
        downloader = MediaIoBaseDownload(buffer, request)
        done = False
        while not done: _, done = downloader.next_chunk()
        buffer.seek(0)
        
        # 2. Load Workbook with Macros preserved
        wb = openpyxl.load_workbook(buffer, keep_vba=True)
        
        # 3. ENSURE 'WELCOME' SHEET EXISTS (Critical for Security)
        if 'Welcome' not in wb.sheetnames:
            ws_welcome = wb.create_sheet('Welcome', 0) # Create at index 0
            ws_welcome['A1'] = "Please Enable Macros to Login and View Data."
            ws_welcome.sheet_state = 'visible'
        else:
            # Ensure it is visible if it already exists
            wb['Welcome'].sheet_state = 'visible'

        # 4. Update the Data Sheet
        if sheet_name_to_update in wb.sheetnames:
            idx = wb.sheetnames.index(sheet_name_to_update)
            wb.remove(wb[sheet_name_to_update])
            ws = wb.create_sheet(sheet_name_to_update, index=idx)
        else:
            ws = wb.create_sheet(sheet_name_to_update)
            
        # Write Data
        for r in dataframe_to_rows(df_to_paste, index=False, header=True):
            ws.append(r)
            
        # 5. LOCK THE DATA SHEET (veryHidden) 
        ws.sheet_state = 'veryHidden' 
            
        # 6. Save and Upload
        out_buffer = BytesIO()
        wb.save(out_buffer)
        out_buffer.seek(0)
        wb.close()
        
        media = MediaIoBaseUpload(out_buffer, mimetype='application/vnd.ms-excel.sheet.macroEnabled.12', resumable=True)
        drive_service.files().update(fileId=file_id, media_body=media, supportsAllDrives=True).execute()
        log("  ✅ XLSM updated and LOCKED successfully.")

    except Exception as e:
        log(f"  [ERROR] Updating XLSM failed: {e}")

def upload_df_as_csv(drive_service, df, file_name, folder_id):
    """Simple helper to upload a DataFrame as a CSV."""
    if df is None: return
    try:
        buffer = io.StringIO()
        df.to_csv(buffer, index=False)
        buffer.seek(0)
        media = MediaIoBaseUpload(BytesIO(buffer.getvalue().encode('utf-8')), mimetype='text/csv', resumable=True)
        
        file_metadata = {'name': file_name, 'parents': [folder_id]}
        drive_service.files().create(body=file_metadata, media_body=media, supportsAllDrives=True).execute()
        log(f"  ✅ CSV Uploaded: {file_name}")
    except Exception as e:
        log(f"  [ERROR] CSV Upload failed: {e}")

# ==============================================================================
# SECTION 2: DATA TRANSFORMATION LOGIC
# ==============================================================================

def process_overall_instock(df):
    """Adds primary keys to Instock Report."""
    if df is None: return None
    try:
        if 'Store Nbr' in df.columns and 'Old Nbr' in df.columns:
            df.insert(0, 'key', df['Store Nbr'].astype(str) + df['Old Nbr'].astype(str))
        return df
    except Exception as e:
        log(f"  [ERROR] Instock processing failed: {e}")
        return None

def process_lmtd_logic(df_lmtd, calc_date):
    """Calculates LMTD & LM Sales."""
    if df_lmtd is None: return None
    log("    > Calculating LMTD & LM Sales from December data...")
    try:
        if 'STORE_NBR' in df_lmtd.columns and 'ITEM_NUMBER' in df_lmtd.columns:
            s_store = pd.to_numeric(df_lmtd['STORE_NBR'], errors='coerce').fillna(0).astype('int64').astype(str)
            s_item = pd.to_numeric(df_lmtd['ITEM_NUMBER'], errors='coerce').fillna(0).astype('int64').astype(str)
            df_lmtd['LMTD_Key'] = s_store + s_item
        else:
            return None

        day_limit = calc_date.day 
        cols_lmtd = []
        for d in range(1, day_limit + 1):
            col_name = f"Sales_December_{d:02d}"
            if col_name in df_lmtd.columns:
                cols_lmtd.append(col_name)
        
        if not cols_lmtd:
            df_lmtd['LMTD Sales'] = 0
        else:
            df_lmtd['LMTD Sales'] = df_lmtd[cols_lmtd].apply(pd.to_numeric, errors='coerce').sum(axis=1)

        cols_lm = [c for c in df_lmtd.columns if c.startswith('Sales_December_')]
        if not cols_lm:
            df_lmtd['LM Sales'] = 0
        else:
            df_lmtd['LM Sales'] = df_lmtd[cols_lm].apply(pd.to_numeric, errors='coerce').sum(axis=1)

        return df_lmtd[['LMTD_Key', 'LMTD Sales', 'LM Sales']]

    except Exception as e:
        log(f"    [ERROR] LMTD Calculation failed: {e}")
        return None

def process_lytd_logic(df_lytd, calc_date):
    """Calculates LYTD & LYM Sales."""
    if df_lytd is None: return None
    log("    > Calculating LYTD & LYM Sales from Jan 2025 data...")
    try:
        if 'STORE_NBR' in df_lytd.columns and 'ITEM_NUMBER' in df_lytd.columns:
            s_store = pd.to_numeric(df_lytd['STORE_NBR'], errors='coerce').fillna(0).astype('int64').astype(str)
            s_item = pd.to_numeric(df_lytd['ITEM_NUMBER'], errors='coerce').fillna(0).astype('int64').astype(str)
            df_lytd['LYTD_Key'] = s_store + s_item
        else:
            return None

        day_limit = calc_date.day 
        cols_lytd = []
        for d in range(1, day_limit + 1):
            col_name = f"Sales_Jan_{d:02d}" 
            if col_name in df_lytd.columns:
                cols_lytd.append(col_name)
        
        if not cols_lytd:
            df_lytd['LYTD Sales'] = 0
        else:
            df_lytd['LYTD Sales'] = df_lytd[cols_lytd].apply(pd.to_numeric, errors='coerce').sum(axis=1)

        cols_lym = [c for c in df_lytd.columns if c.startswith('Sales_Jan_')]
        if not cols_lym:
            df_lytd['LYM Sales'] = 0
        else:
            df_lytd['LYM Sales'] = df_lytd[cols_lym].apply(pd.to_numeric, errors='coerce').sum(axis=1)

        return df_lytd[['LYTD_Key', 'LYTD Sales', 'LYM Sales']]

    except Exception as e:
        log(f"    [ERROR] LYTD Calculation failed: {e}")
        return None

def process_article_sales_report(df, df_hirarchy, df_div, df_instock, df_gst, df_ytd, df_lmtd, df_lytd, calc_date):
    """The Master Transformation Function."""
    log("    > Processing Article Sales Report (Transformation Pipeline)...")
    if df is None: return None
    
    try:
        # 1. Generate Article UID
        if 'Article No' in df.columns and 'Store No' in df.columns:
            s_store = pd.to_numeric(df['Store No'], errors='coerce').fillna(0).astype('int64').astype(str)
            s_article = pd.to_numeric(df['Article No'], errors='coerce').fillna(0).astype('int64').astype(str)
            df.insert(df.columns.get_loc('Article No')+1, 'Article UID', s_store + s_article)

        # ======================================================================
        # 2. Merge Hierarchy (ROBUST FIX FOR TEXT vs NUMBER MISMATCH)
        # ======================================================================
        if df_hirarchy is not None and 'Store No' in df.columns:
            # A. Clean Main Data Store No: Force Float -> Int -> String
            # This handles "1001" and "1001.0" and 1001 by forcing them all to "1001"
            df['Store_Key'] = pd.to_numeric(df['Store No'], errors='coerce').fillna(0).astype('int64').astype(str)
            
            # B. Clean Hierarchy Location: Force Float -> Int -> String
            # NOTE: We assume 'Location' is the store number in the hierarchy file
            df_hirarchy['Location_Key'] = pd.to_numeric(df_hirarchy['Location'], errors='coerce').fillna(0).astype('int64').astype(str)
            
            # C. Check if 'Market' or 'Region' exists
            # Sometimes CSV headers vary. We check for both.
            hier_cols = ['Location_Key', 'Market Manager']
            if 'Region' in df_hirarchy.columns:
                hier_cols.append('Region')
            elif 'Market' in df_hirarchy.columns:
                df_hirarchy.rename(columns={'Market': 'Region'}, inplace=True)
                hier_cols.append('Region')
            
            # D. Merge using the clean keys
            df = df.merge(df_hirarchy[hier_cols], left_on='Store_Key', right_on='Location_Key', how='left')
            
            # E. Drop temporary keys
            df.drop(columns=['Location_Key', 'Store_Key'], inplace=True, errors='ignore')

        # 3. Merge Division V1
        if df_div is not None and 'Sub Division' in df.columns:
            df['Sub Division'] = df['Sub Division'].astype(str)
            df_div['Sub Division'] = df_div['Sub Division'].astype(str)
            df = df.merge(df_div[['Sub Division', 'Sub Division_V1']], on='Sub Division', how='left')

        # 4. Merge KVI Status
        if df_instock is not None and 'Article UID' in df.columns:
            df['Article UID'] = df['Article UID'].astype(str)
            df_instock['key'] = df_instock['key'].astype(str)
            df = df.merge(df_instock[['key', 'KVI_Flag', 'KVI_Allocation', 'KVI_Utilization']], left_on='Article UID', right_on='key', how='left')
            df.drop(columns=['key'], inplace=True, errors='ignore')

        # 5. Merge GST Changes
        if df_gst is not None and 'Article UID' in df.columns:
            df_gst['UID'] = df_gst['UID'].astype(str)
            df_gst['GST_Change'] = 'Yes'
            df = df.merge(df_gst[['UID', 'GST_Change']], left_on='Article UID', right_on='UID', how='left')
            df['GST_Change'] = df['GST_Change'].fillna('')
            df.drop(columns=['UID'], inplace=True, errors='ignore')

        # 6. Merge Historical YTD Sales
        if df_ytd is not None and 'Article UID' in df.columns:
            df['Article UID'] = df['Article UID'].astype(str).str.replace(r'\.0$', '', regex=True)
            df_ytd['Article UID'] = pd.to_numeric(df_ytd['Article UID'], errors='coerce').fillna(-1).astype('int64').astype(str)
            
            ytd_cols = [c for c in ['Article UID', '2021 YTD Sales', '2022 YTD Sales', '2023 YTD Sales', '2024 YTD Sales', '2025 YTD Sales'] if c in df_ytd.columns]
            df = df.merge(df_ytd[ytd_cols], on='Article UID', how='left')

        # 7. Merge LMTD
        if df_lmtd is not None and 'Article UID' in df.columns:
             df = df.merge(df_lmtd, left_on='Article UID', right_on='LMTD_Key', how='left')
             df['LMTD Sales'] = df['LMTD Sales'].fillna(0)
             df['LM Sales'] = df['LM Sales'].fillna(0)
             df.drop(columns=['LMTD_Key'], inplace=True, errors='ignore')

        # 8. Merge LYTD
        if df_lytd is not None and 'Article UID' in df.columns:
             df = df.merge(df_lytd, left_on='Article UID', right_on='LYTD_Key', how='left')
             df['LYTD Sales'] = df['LYTD Sales'].fillna(0)
             df['LYM Sales'] = df['LYM Sales'].fillna(0)
             df.drop(columns=['LYTD_Key'], inplace=True, errors='ignore')

        # 9. Calculate Metrics
        day_of_year = calc_date.timetuple().tm_yday
        if day_of_year > 0:
            sales_cols = ['YTD Sale Amt', '2021 YTD Sales', '2022 YTD Sales', '2023 YTD Sales', '2024 YTD Sales', '2025 YTD Sales']
            for col in sales_cols:
                if col in df.columns:
                    avg_col_name = col.replace('YTD Sales', 'Avg Sales').replace('Sale Amt', 'Avg Sales')
                    df[avg_col_name] = pd.to_numeric(df[col], errors='coerce') / day_of_year
            
            if 'YTD COST Amt' in df.columns and 'On Hand Cost' in df.columns:
                daily_cost_burn = pd.to_numeric(df['YTD COST Amt'], errors='coerce') / day_of_year
                on_hand_val = pd.to_numeric(df['On Hand Cost'], errors='coerce')
                df['Day On Hand'] = on_hand_val / daily_cost_burn
                df['Day On Hand'] = df['Day On Hand'].replace([np.inf, -np.inf], np.nan)
                
                conditions = [(df['Day On Hand'] > 7), (df['Day On Hand'] <= 7)]
                choices = ['Price Support Required', 'Stock Required']
                df['Final Remarks'] = np.select(conditions, choices, default='')

        # 10. Clean up
        cols_drop = ['WEEK4_COST', 'WEEK4_QTY', 'WEEEK4_Sales', 'WEEK4_Sales']
        df.drop(columns=[c for c in cols_drop if c in df.columns], inplace=True)

        if 'Article Status' in df.columns: 
            df = df[df['Article Status'].astype(str).str.strip().str.upper() != 'D']
        if 'Division' in df.columns: 
            df = df[~df['Division'].astype(str).str.lower().isin(['freebies', 'service article'])]
        if 'Store' in df.columns: 
            df = df[df['Store'].astype(str).str.strip().str.lower() != 'lucknow fc']

        # 11. REORDER COLUMNS
        desired_order = [
            "Article No", "Article UID", "Store No", "Store", "Region", "Market Manager",
            "Article Description", "Brand Name", "Article Type", "PB_FLAG", "Base Unit of Measurement",
            "RP Type", "Article Status", "Purchase Group", "Division", "Sub Division", "Sub Division_V1",
            "Category", "Sub Category No", "Sub Category", "Fineline No", "Fineline", "Vendor No",
            "Vendor Name", "Last GRN Date", "VNPK Qty", "VNPK Cost", "KVI_Flag", "KVI_Allocation",
            "KVI_Utilization", "MAP / WHPK", "Selling Price (With Tax)", "Selling Price (Without Tax)",
            "Current MRP", "On Hand Qty", "On Hand Cost", "On Order Qty", "On Order Cost",
            "FTD Qty", "FTD Sale Amt", "FTD COST Amt", "FTD IM %", "MTD Qty", "MTD Sale Amt",
            "MTD COST Amt", "MTD IM %", "YTD Qty", "YTD Sale Amt", "YTD COST Amt", "YTD IM %",
            "GST_Change", "2021 YTD Sales", "2022 YTD Sales", "2023 YTD Sales", "2024 YTD Sales",
            "2025 YTD Sales", "YTD Avg Sales", "2021 Avg Sales", "2022 Avg Sales", "2023 Avg Sales",
            "2024 Avg Sales", "2025 Avg Sales", "Day On Hand", 
            "LMTD Sales", "LM Sales", "LYTD Sales", "LYM Sales", "Final Remarks"
        ]
        
        final_columns = [col for col in desired_order if col in df.columns]
        remaining_cols = [col for col in df.columns if col not in final_columns]
        final_columns.extend(remaining_cols)
        
        df = df[final_columns]
        log(f"    > Columns Reordered. Final Rows: {len(df)}")
        return df

    except Exception as e:
        log(f"    [ERROR] Transformation failed: {e}")
        return None

# ==============================================================================
# SECTION 3: ADVANCED INTELLIGENCE ENGINE (EXCEL + GRAPHS)
# ==============================================================================

def generate_excel_insights_report(df, date_str):
    """Generates Excel Dashboard with specific Store-Level Opportunity Analysis."""
    log("    > Spinning up Intelligence Engine (Excel Generation)...")
    
    output = BytesIO()
    
    # Ensure numeric columns are actually numeric
    numeric_cols = [
        'YTD Sale Amt', 'FTD Sale Amt', 'On Hand Cost', 'YTD Avg Sales', 
        'Day On Hand', 'MTD Sale Amt', 'LMTD Sales', 'LYTD Sales',
        'LM Sales', 'LYM Sales', 'On Hand Qty', 'Selling Price (With Tax)'
    ]
    for col in numeric_cols:
        if col in df.columns: 
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    with pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}}) as writer:
        workbook = writer.book
        
        # --- FORMATS ---
        fmt_header = workbook.add_format({'bold': True, 'bg_color': '#2F75B5', 'font_color': 'white', 'border': 1})
        fmt_subhead = workbook.add_format({'bold': True, 'bg_color': '#DDEBF7', 'border': 1})
        fmt_currency = workbook.add_format({'num_format': '₹ #,##0.00'})
        fmt_number = workbook.add_format({'num_format': '#,##0'})
        fmt_pct = workbook.add_format({'num_format': '0.0%'})
        
        # ======================================================================
        # SHEET 1: DASHBOARD
        # ======================================================================
        ws_dash = workbook.add_worksheet('Dashboard')
        
        total_ftd = df['FTD Sale Amt'].sum() if 'FTD Sale Amt' in df.columns else 0
        total_ytd = df['YTD Sale Amt'].sum() if 'YTD Sale Amt' in df.columns else 0
        total_inv = df['On Hand Cost'].sum() if 'On Hand Cost' in df.columns else 0
        total_lmtd = df['LMTD Sales'].sum() if 'LMTD Sales' in df.columns else 0
        total_lytd = df['LYTD Sales'].sum() if 'LYTD Sales' in df.columns else 0
        efficiency = (total_ytd / total_inv) if total_inv > 0 else 0
        
        data_kpi = [
            ['KPI Metric', 'Value'],
            ['Total FTD Sales (Today)', total_ftd],
            ['Total YTD Sales', total_ytd],
            ['Total LMTD Sales', total_lmtd],
            ['Total LYTD Sales', total_lytd],
            ['Total Inventory Value', total_inv],
            ['Capital Efficiency Ratio', efficiency]
        ]
        
        ws_dash.write('B2', f"DAILY INTELLIGENCE REPORT: {date_str}", workbook.add_format({'bold': True, 'font_size': 14}))
        
        for i, row in enumerate(data_kpi):
            fmt = fmt_header if i == 0 else fmt_currency
            if i == 6: fmt = workbook.add_format({'num_format': '0.00'}) 
            ws_dash.write(4 + i, 1, row[0], fmt)
            ws_dash.write(4 + i, 2, row[1], fmt)

        if 'Store' in df.columns and 'FTD Sale Amt' in df.columns:
            top_stores = df.groupby('Store')['FTD Sale Amt'].sum().nlargest(10).reset_index()
            top_stores.to_excel(writer, sheet_name='Dashboard', startrow=12, startcol=1, index=False)
            
            chart_col = workbook.add_chart({'type': 'column'})
            chart_col.add_series({
                'name': 'FTD Sales',
                'categories': ['Dashboard', 13, 1, 22, 1],
                'values':     ['Dashboard', 13, 2, 22, 2],
                'fill':       {'color': '#4472C4'}
            })
            chart_col.set_title({'name': 'Top 10 Stores (Today\'s Sales)'})
            ws_dash.insert_chart('E5', chart_col)

        # ======================================================================
        # SHEET 2: STORE-LEVEL PARETO
        # ======================================================================
        ws_pareto = workbook.add_worksheet('Pareto_Analysis')
        
        if 'Store' in df.columns and 'FTD Sale Amt' in df.columns and 'Article Description' in df.columns:
            ws_pareto.write('A1', 'STORE PARETO SUMMARY (Based on FTD Sales)', fmt_header)
            headers = ['Store', 'Total Articles Sold', 'Power SKUs (80% Sales)', 'Tail SKUs (20% Sales)', 
                       'Total FTD Sales', 'Total MTD Sales', 'Total YTD Sales', 'Total LMTD Sales', 'Total LYTD Sales']
            for c, h in enumerate(headers): ws_pareto.write(2, c, h, fmt_subhead)
            
            stores = df['Store'].unique()
            row_idx = 3
            pareto_details_list = []
            
            for store in stores:
                store_df = df[(df['Store'] == store) & (df['FTD Sale Amt'] > 0)].copy()
                if store_df.empty: continue
                
                store_df = store_df.sort_values(by='FTD Sale Amt', ascending=False)
                total_sales = store_df['FTD Sale Amt'].sum()
                store_df['Cum_Sales'] = store_df['FTD Sale Amt'].cumsum()
                store_df['Cum_Pct'] = store_df['Cum_Sales'] / total_sales
                
                store_df['Pareto_Class'] = np.where(store_df['Cum_Pct'] <= 0.80, 'A (Top 80%)', 'B (Tail 20%)')
                
                total_articles = len(store_df)
                power_skus = len(store_df[store_df['Pareto_Class'] == 'A (Top 80%)'])
                tail_skus = total_articles - power_skus
                
                ftd_val = store_df['FTD Sale Amt'].sum()
                mtd_val = store_df['MTD Sale Amt'].sum()
                ytd_val = store_df['YTD Sale Amt'].sum()
                lmtd_val = store_df['LMTD Sales'].sum() if 'LMTD Sales' in store_df.columns else 0
                lytd_val = store_df['LYTD Sales'].sum() if 'LYTD Sales' in store_df.columns else 0
                
                ws_pareto.write(row_idx, 0, store)
                ws_pareto.write(row_idx, 1, total_articles, fmt_number)
                ws_pareto.write(row_idx, 2, power_skus, fmt_number)
                ws_pareto.write(row_idx, 3, tail_skus, fmt_number)
                ws_pareto.write(row_idx, 4, ftd_val, fmt_currency)
                ws_pareto.write(row_idx, 5, mtd_val, fmt_currency)
                ws_pareto.write(row_idx, 6, ytd_val, fmt_currency)
                ws_pareto.write(row_idx, 7, lmtd_val, fmt_currency)
                ws_pareto.write(row_idx, 8, lytd_val, fmt_currency)
                
                row_idx += 1
                
                power_items = store_df[store_df['Pareto_Class'] == 'A (Top 80%)'][['Store', 'Article UID', 'Article Description', 'FTD Sale Amt', 'Pareto_Class']]
                pareto_details_list.append(power_items)

            detail_start_row = row_idx + 3
            ws_pareto.write(detail_start_row, 0, 'DETAILED POWER SKUs (Articles contributing to 80% of Sales today)', fmt_header)
            
            if pareto_details_list:
                full_pareto_df = pd.concat(pareto_details_list)
                full_pareto_df.to_excel(writer, sheet_name='Pareto_Analysis', startrow=detail_start_row+1, index=False)

        # ======================================================================
        # SHEET 3: REGIONAL DEEP DIVE (MODIFIED - Grouped + MTD Sales/Stock)
        # ======================================================================
        if 'Store' in df.columns:
            ws_region = workbook.add_worksheet('Regional_Deep_Dive')
            row_cursor = 0
            
            store_groups = {
                'Group-1 (AP Coastal)': ['Guntur', 'Vijayawada', 'Rajahmundry', 'Visakhapatnam'],
                'Group-2 (Telangana/Rayalaseema)': ['Karimnagar', 'Hyderabad', 'Kurnool', 'Tirupathi'],
                'Group-3 (Central)': ['Indore-1', 'Indore-2', 'Bhopal-1', 'Bhopal-2', 'Aurangabad', 'Amravati', 'Raipur'],
                'Group-4 (UP/North)': ['Agra-1', 'Agra-2', 'Lucknow', 'Meerut', 'Kota'],
                'Group-5 (Punjab/J&K)': ['Amritsar', 'Jammu', 'Ludhiana-1', 'Ludhiana-3', 'Jalandhar', 'Zirakpur']
            }

            for group_name, store_list in store_groups.items():
                # Filter Main DF for stores in this group
                group_df = df[df['Store'].astype(str).str.strip().isin(store_list)].copy()
                if group_df.empty: continue

                # WRITE GROUP HEADER
                ws_region.merge_range(row_cursor, 0, row_cursor, len(store_list), f"--- {group_name} ---", fmt_header)
                row_cursor += 1

                # MTD Sales & Stock Logic
                if 'MTD Sale Amt' in group_df.columns:
                    top_arts = group_df.groupby('Article Description')['MTD Sale Amt'].sum().nlargest(10).index.tolist()
                    subset = group_df[group_df['Article Description'].isin(top_arts)]
                    
                    pivot_sales = subset.pivot_table(index='Article Description', columns='Store', values='MTD Sale Amt', aggfunc='sum').fillna(0)
                    pivot_stock = subset.pivot_table(index='Article Description', columns='Store', values='On Hand Qty', aggfunc='sum').fillna(0)
                    
                    # Ensure column order matches group list
                    cols_present = [s for s in store_list if s in pivot_sales.columns]
                    pivot_sales = pivot_sales[cols_present]
                    pivot_stock = pivot_stock[cols_present]
                    
                    # Table 1: MTD Sales
                    ws_region.write(row_cursor, 0, "Top 10 Articles (By MTD Sales)", fmt_subhead)
                    for c, col in enumerate(cols_present): ws_region.write(row_cursor, c+1, col, fmt_subhead)
                    row_cursor += 1
                    
                    for art in top_arts:
                        if art in pivot_sales.index:
                            ws_region.write(row_cursor, 0, art)
                            for c, val in enumerate(pivot_sales.loc[art]):
                                ws_region.write(row_cursor, c+1, val, fmt_currency)
                            row_cursor += 1
                    
                    row_cursor += 1 # Spacer

                    # Table 2: On Hand Qty
                    ws_region.write(row_cursor, 0, "Stock Status (On Hand Qty)", fmt_subhead)
                    for c, col in enumerate(cols_present): ws_region.write(row_cursor, c+1, col, fmt_subhead)
                    row_cursor += 1
                    
                    for art in top_arts:
                        if art in pivot_stock.index:
                            ws_region.write(row_cursor, 0, art)
                            for c, val in enumerate(pivot_stock.loc[art]):
                                ws_region.write(row_cursor, c+1, val, fmt_number)
                            row_cursor += 1
                    
                    row_cursor += 3

        # ======================================================================
        # SHEET 4: ACTIONABLES
        # ======================================================================
        ws_action = workbook.add_worksheet('Actionables')
        ws_action.write('A1', 'URGENT REORDER (Top 50 Fast Movers)', fmt_header)
        if 'Final Remarks' in df.columns:
            urgent = df[(df['Final Remarks'] == 'Stock Required') & (df['YTD Avg Sales'] > 0)].sort_values('YTD Avg Sales', ascending=False).head(50)
            cols_urg = ['Article UID', 'Article Description', 'Store', 'Day On Hand', 'YTD Avg Sales']
            for c, col in enumerate(cols_urg): ws_action.write(1, c, col, fmt_subhead)
            for r, row in enumerate(urgent[cols_urg].values):
                for c, val in enumerate(row):
                    ws_action.write(r+2, c, val)

        ws_action.write('G1', 'CASH TRAPS (High Value Dead Stock)', fmt_header)
        if 'Day On Hand' in df.columns:
            traps = df[(df['Day On Hand'] > 180) & (df['On Hand Cost'] > 50000)].sort_values('On Hand Cost', ascending=False).head(50)
            cols_trap = ['Article UID', 'Article Description', 'Store', 'Day On Hand', 'On Hand Cost']
            for c, col in enumerate(cols_trap): ws_action.write(1, c+6, col, fmt_subhead)
            for r, row in enumerate(traps[cols_trap].values):
                for c, val in enumerate(row):
                    ws_action.write(r+2, c+6, val)

        ws_action.write('M1', 'MARGIN BLEED (Category Level)', fmt_header)
        cat_col = 'Sub Division_V1' if 'Sub Division_V1' in df.columns else 'Sub Division'
        if cat_col in df.columns and 'MTD IM %' in df.columns:
            margin = df.groupby(cat_col)[['MTD IM %', 'YTD IM %']].mean()
            margin['Drop'] = margin['YTD IM %'] - margin['MTD IM %']
            bleeders = margin[margin['Drop'] > 2].sort_values('Drop', ascending=False).reset_index()
            cols_marg = [cat_col, 'MTD IM %', 'YTD IM %', 'Drop']
            for c, col in enumerate(cols_marg): ws_action.write(1, c+12, col, fmt_subhead)
            for r, row in enumerate(bleeders[cols_marg].values):
                for c, val in enumerate(row):
                    ws_action.write(r+2, c+12, val)

        # ======================================================================
        # SHEET 5: CORRELATIONS
        # ======================================================================
        ws_corr = workbook.add_worksheet('Correlations')
        
        ws_corr.write('A1', 'TOP VENDORS CAUSING STOCKOUTS', fmt_header)
        if 'Vendor Name' in df.columns:
            vendor_risk = df[df['Final Remarks'] == 'Stock Required'].groupby('Vendor Name')['Article UID'].count().reset_index(name='Stockout Count')
            vendor_risk = vendor_risk.sort_values('Stockout Count', ascending=False).head(20)
            
            ws_corr.write(1, 0, 'Vendor Name')
            ws_corr.write(1, 1, 'Stockout Count')
            for r, row in enumerate(vendor_risk.values):
                ws_corr.write(r+2, 0, row[0])
                ws_corr.write(r+2, 1, row[1])

        ws_corr.write('E1', 'GST PRICE IMPACT (Elasticity Proxy)', fmt_header)
        if 'GST_Change' in df.columns and '2024 Avg Sales' in df.columns and 'MTD Qty' in df.columns:
            gst_items = df[df['GST_Change'] == 'Yes'].copy()
            if not gst_items.empty and 'Selling Price (With Tax)' in df.columns:
                gst_items['Est_Pre_Vol'] = gst_items['2024 Avg Sales'] / gst_items['Selling Price (With Tax)'].replace(0, 1)
                gst_items['Est_Post_Vol'] = gst_items['MTD Qty'] / 30
                gst_items['Vol Change %'] = ((gst_items['Est_Post_Vol'] - gst_items['Est_Pre_Vol']) / gst_items['Est_Pre_Vol'].replace(0, 1))
                gst_items['Vol Change %'] = gst_items['Vol Change %'].fillna(0)
                
                view = gst_items[['Article Description', 'Selling Price (With Tax)', 'Vol Change %']].head(50)
                
                cols_gst = ['Article Description', 'Price', 'Vol Change %']
                for c, col in enumerate(cols_gst): ws_corr.write(1, c+4, col)
                for r, row in enumerate(view.values):
                    ws_corr.write(r+2, 4, row[0])
                    ws_corr.write(r+2, 5, row[1])
                    ws_corr.write(r+2, 6, row[2], fmt_pct)

        # ======================================================================
        # SHEET 6: OPPORTUNITY ANALYSIS (NEW - High LMTD vs Low MTD + High Stock)
        # ======================================================================
        ws_opp = workbook.add_worksheet('Opp_HighLMTD_LowMTD')
        
        req_cols = ['Store', 'Article Description', 'LMTD Sales', 'MTD Sale Amt', 'On Hand Qty', 'LYM Sales', 'Selling Price (With Tax)']
        if all(c in df.columns for c in req_cols):
            
            # 1. Calculate the Drop (Dip)
            df['Sales_Drop_Value'] = df['LMTD Sales'] - df['MTD Sale Amt']
            
            # 2. Filter: Drop > 0 (Sales declined) AND Stock > 100
            mask = (df['Sales_Drop_Value'] > 0) & (df['On Hand Qty'] > 100)
            opp_df = df[mask].copy()
            
            # 3. Sort by Store (Asc) and Drop Value (Desc)
            opp_df = opp_df.sort_values(by=['Store', 'Sales_Drop_Value'], ascending=[True, False])
            
            # 4. Get Top 10 per Store
            top_opps = opp_df.groupby('Store').head(10)
            
            # 5. Define Output Columns
            out_cols = [
                'Store', 'Article Description', 
                'LMTD Sales', 'MTD Sale Amt', 'Sales_Drop_Value', 
                'On Hand Qty', 'LYM Sales', 'Selling Price (With Tax)'
            ]
            
            # 6. Write to Excel
            ws_opp.write('A1', "Opportunity Analysis: High LMTD vs Low MTD (Stock > 100)", fmt_header)
            ws_opp.write('A2', "Top 10 articles per store where sales dropped significantly despite high stock.", workbook.add_format({'italic': True}))
            
            # Write Headers
            for c, col in enumerate(out_cols):
                ws_opp.write(3, c, col, fmt_subhead)
                
            # Write Rows
            for r, row in enumerate(top_opps[out_cols].values):
                row_num = r + 4
                ws_opp.write(row_num, 0, row[0]) # Store
                ws_opp.write(row_num, 1, row[1]) # Desc
                ws_opp.write(row_num, 2, row[2], fmt_currency) # LMTD
                ws_opp.write(row_num, 3, row[3], fmt_currency) # MTD
                ws_opp.write(row_num, 4, row[4], fmt_currency) # Drop Value
                ws_opp.write(row_num, 5, row[5], fmt_number)   # On Hand Qty
                ws_opp.write(row_num, 6, row[6], fmt_currency) # LYM
                ws_opp.write(row_num, 7, row[7], fmt_currency) # Price
            
            # 7. ADD PLOTS (Conditional Formatting / Data Bars)
            if not top_opps.empty:
                last_row = len(top_opps) + 4
                
                # Plot 1: Red Data Bar for Sales Drop (Column E / Index 4)
                ws_opp.conditional_format(4, 4, last_row, 4, {
                    'type': 'data_bar', 
                    'bar_color': '#FF6347', # Tomato Red
                    'bar_solid': True
                })
                
                # Plot 2: Blue/Green Data Bar for On Hand Qty (Column F / Index 5)
                ws_opp.conditional_format(4, 5, last_row, 5, {
                    'type': 'data_bar', 
                    'bar_color': '#63C384', # Green
                    'bar_solid': True
                })
            
            # Set Column Widths for readability
            ws_opp.set_column(0, 0, 15) # Store
            ws_opp.set_column(1, 1, 40) # Desc
            ws_opp.set_column(2, 7, 12) # Metrics

    output.seek(0)
    return output

# ==============================================================================
# SECTION 4: MAIN ORCHESTRATOR
# ==============================================================================

def find_files_for_date(drive_service, date_str):
    """Finds source CSV/ZIP files for a given date."""
    log(f"  Querying source files for date: {date_str}")
    file_info = {}
    for prefix in FILE_PREFIXES:
        q = f"'{SOURCE_FOLDER_ID}' in parents and (name='{prefix}_{date_str}.csv' or name='{prefix}_{date_str}.zip') and trashed=false"
        results = drive_service.files().list(q=q, fields="files(id, name)").execute()
        items = results.get('files', [])
        if not items: return None
        file_info[prefix] = (items[0]['id'], items[0]['name'])
    return file_info

def check_and_copy_files(drive_service):
    """Main execution function."""
    if not drive_service: return

    # 1. Download Helpers
    log("\n--- Phase 1: Downloading Reference Data ---")
    df_hirarchy = download_csv_to_df(drive_service, 'hirarchy.csv', TARGET_FOLDER_ID)
    df_div = download_csv_to_df(drive_service, 'division_group.csv', TARGET_FOLDER_ID)
    df_gst = download_csv_to_df(drive_service, 'gst_change_list.csv', TARGET_FOLDER_ID)
    df_ytd = download_csv_to_df(drive_service, 'ytd_sales.csv', TARGET_FOLDER_ID)
    
    df_lmtd_raw = download_csv_to_df(drive_service, '2025_dec_sales.csv', TARGET_FOLDER_ID)
    df_lytd_raw = download_csv_to_df(drive_service, '2025_jan_sales.csv', TARGET_FOLDER_ID)

    # 2. Date Fallback Logic
    log("\n--- Phase 2: Locating Source Files ---")
    today = datetime.date.today()
    date_str = today.strftime('%Y-%m-%d')
    calc_date = today - datetime.timedelta(days=1)
    
    file_info = find_files_for_date(drive_service, date_str)
    
    if not file_info:
        log(f"  [WARN] Files for {date_str} not found. Checking yesterday...")
        date_str = (today - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        calc_date = today - datetime.timedelta(days=2)
        file_info = find_files_for_date(drive_service, date_str)

    if not file_info:
        log("❌ CRITICAL: No source files found for Today or Yesterday. Exiting.")
        return

    log(f"✅ Found files for {date_str}. Proceeding...")
    
    # 3. Load & Process Main Files
    log("\n--- Phase 3: Loading & Processing ---")
    df_article = load_file_to_df(drive_service, file_info['ArticleSalesReport'][0], file_info['ArticleSalesReport'][1])
    df_instock = load_file_to_df(drive_service, file_info['Overall_Instock'][0], file_info['Overall_Instock'][1])
    
    df_instock = process_overall_instock(df_instock)
    df_lmtd_clean = process_lmtd_logic(df_lmtd_raw, calc_date)
    df_lytd_clean = process_lytd_logic(df_lytd_raw, calc_date)
    
    df_final = process_article_sales_report(
        df_article, df_hirarchy, df_div, df_instock, df_gst, df_ytd, 
        df_lmtd_clean, df_lytd_clean, calc_date
    )
    
    # 4. Generate Outputs
    if df_final is not None:
        log("\n--- Phase 4: Output Generation ---")
        
        # A. Update Raw Data XLSM (SECURE VERSION)
        update_xlsm_data_sheet(
            drive_service, df_final, 
            "article_sales_report.xlsm", "Sheet1", TARGET_FOLDER_ID
        )
        
        # B. Upload Raw Data CSV
        if df_instock is not None:
            upload_df_as_csv(drive_service, df_instock, f"Overall_Instock_{date_str}.csv", TARGET_FOLDER_ID)

        # C. Generate & Upload Advanced Excel Dashboard
        excel_buffer = generate_excel_insights_report(df_final, date_str)
        upload_excel_report(drive_service, excel_buffer, f"Business_Insights_Report_{date_str}.xlsx", TARGET_FOLDER_ID)

    # 5. Copy Originals & Cleanup
    log("\n--- Phase 5: Backup Original Files ---")
    for prefix in FILE_PREFIXES:
        f_id, f_name = file_info[prefix]
        q = f"'{TARGET_FOLDER_ID}' in parents and name='{f_name}' and trashed=false"
        if not drive_service.files().list(q=q).execute().get('files'):
            drive_service.files().copy(fileId=f_id, body={'name': f_name, 'parents': [TARGET_FOLDER_ID]}).execute()
            
    log("  Cleaning up memory...")
    del df_article, df_instock, df_final, df_hirarchy, df_div, df_gst, df_ytd, df_lmtd_raw, df_lytd_raw
    gc.collect()
    
    log("\n=== SUCCESS: Pipeline Completed Successfully ===")

if __name__ == "__main__":
    srv = authenticate()
    if srv: check_and_copy_files(srv)
